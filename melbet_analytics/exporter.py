import json
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import REPORTS_DIR, logger

def export_to_json(user_data: dict, metrics: dict, summaries: dict, filename: str = None) -> str:
    """Exports all scraped and analyzed data to a JSON file."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    user_id = user_data.get("user_id", "UNKNOWN")
    
    if not filename:
        filename = f"melbet_report_{user_id}_{timestamp}.json"
        
    out_path = REPORTS_DIR / filename
    
    report_content = {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "target": "Melbet Account Analytics System"
        },
        "user_profile": {
            "user_id": user_id,
            "username": user_data.get("username", "N/A"),
            "wallet_balance": user_data.get("wallet_balance", 0.0),
            "currency": user_data.get("currency", "INR"),
            "account_status": user_data.get("account_status", "Active")
        },
        "overall_metrics": metrics,
        "periodic_summaries": summaries,
        "deposits": user_data.get("deposits", []),
        "withdrawals": user_data.get("withdrawals", []),
        "bets": user_data.get("bets", [])
    }
    
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(report_content, f, indent=4, ensure_ascii=False)
        logger.info(f"JSON report exported to: {out_path}")
    except Exception as e:
        logger.error(f"Failed to export JSON report: {e}", exc_info=True)
        
    return str(out_path)

def export_to_csv(user_data: dict, user_id: str):
    """Exports raw deposits, withdrawals, and bets lists to individual CSV files."""
    try:
        for key in ["deposits", "withdrawals", "bets"]:
            records = user_data.get(key, [])
            if records:
                df = pd.DataFrame(records)
                csv_path = REPORTS_DIR / f"{key}_{user_id}.csv"
                df.to_csv(csv_path, index=False, encoding="utf-8")
                logger.info(f"CSV export of {key} saved to: {csv_path}")
            else:
                logger.info(f"No {key} records found to export to CSV.")
    except Exception as e:
        logger.error(f"Failed to export CSVs: {e}", exc_info=True)

def export_to_excel(user_data: dict, metrics: dict, summaries: dict, filename: str = None) -> str:
    """Exports data into a professional, styled, multi-sheet Excel file."""
    user_id = user_data.get("user_id", "UNKNOWN")
    
    if not filename:
        filename = f"melbet_analytics_{user_id}.xlsx"
        
    out_path = REPORTS_DIR / filename
    
    try:
        # Create DataFrames
        df_deps = pd.DataFrame(user_data.get("deposits", []))
        df_wds = pd.DataFrame(user_data.get("withdrawals", []))
        df_bets = pd.DataFrame(user_data.get("bets", []))
        
        # Summary Overview Data
        summary_records = [
            {"Metric": "User ID", "Value": user_id},
            {"Metric": "Username", "Value": user_data.get("username", "N/A")},
            {"Metric": "Wallet Balance", "Value": f"{user_data.get('wallet_balance', 0.0):,.2f} {user_data.get('currency', 'INR')}"},
            {"Metric": "Account Status", "Value": user_data.get("account_status", "Active")},
            {"Metric": "Total Deposits", "Value": f"{metrics['total_deposits']:,.2f}"},
            {"Metric": "Total Withdrawals", "Value": f"{metrics['total_withdrawals']:,.2f}"},
            {"Metric": "Total Bets Placed", "Value": metrics["total_bets"]},
            {"Metric": "Total Wins", "Value": metrics["total_wins"]},
            {"Metric": "Total Losses", "Value": metrics["total_losses"]},
            {"Metric": "Total Stake", "Value": f"{metrics['total_stake']:,.2f}"},
            {"Metric": "Net Profit/Loss", "Value": f"{metrics['net_pnl']:,.2f}"},
            {"Metric": "Return on Investment (ROI %)", "Value": f"{metrics['roi']}%"}
        ]
        df_summary = pd.DataFrame(summary_records)
        
        # Periodic summaries
        df_daily = pd.DataFrame(summaries.get("daily", []))
        df_weekly = pd.DataFrame(summaries.get("weekly", []))
        df_monthly = pd.DataFrame(summaries.get("monthly", []))
        
        # Write sheets to Excel writer
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Overview", index=False)
            
            if not df_deps.empty:
                df_deps.to_excel(writer, sheet_name="Deposits", index=False)
            if not df_wds.empty:
                df_wds.to_excel(writer, sheet_name="Withdrawals", index=False)
            if not df_bets.empty:
                df_bets.to_excel(writer, sheet_name="Bets", index=False)
                
            if not df_daily.empty:
                df_daily.to_excel(writer, sheet_name="Daily Summary", index=False)
            if not df_weekly.empty:
                df_weekly.to_excel(writer, sheet_name="Weekly Summary", index=False)
            if not df_monthly.empty:
                df_monthly.to_excel(writer, sheet_name="Monthly Summary", index=False)
                
        # --- STYLING WORKBOOKS USING OPENPYXL ---
        # Reloading Excel file to apply specific styling
        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        
        # Colors & Fonts
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=11)
        bold_font = Font(name="Segoe UI", size=11, bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='BDC3C7'),
            right=Side(style='thin', color='BDC3C7'),
            top=Side(style='thin', color='BDC3C7'),
            bottom=Side(style='thin', color='BDC3C7')
        )
        
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            
            # Formats row headers
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Formatting cells
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = regular_font
                    cell.border = thin_border
                    
                    # Highlight overview values or metric names
                    if ws_name == "Overview" and col == 1:
                        cell.font = bold_font
                    
                    # Standardize alignments
                    if col == 1 and ws_name == "Overview":
                        cell.alignment = Alignment(horizontal="left")
                    elif isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="center")
            
            # Auto-fit columns
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            # Set grid lines visible
            ws.views.sheetView[0].showGridLines = True
            
        wb.save(out_path)
        logger.info(f"Excel analytics workbook generated and formatted at: {out_path}")
        
    except Exception as e:
        logger.error(f"Failed to export Excel report: {e}", exc_info=True)
        
    return str(out_path)
